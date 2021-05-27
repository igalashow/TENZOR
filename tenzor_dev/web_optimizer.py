import requests
from datetime import datetime, date, time
from bs4 import BeautifulSoup
import time
from fake_useragent import UserAgent
import random
import re
import os
from openpyxl import Workbook, load_workbook
import telebot

reg = '32'				# указать в соответствии с регионом
vecher = 22				# время фиксации прогноза (уже вечер, он не изменится. GMT+3)
url = 'http://www.kpl32.ru/?page=37&is_type=13'

chatid = '-1001441037466'
api_token = '1301173613:AAHpTOmYPAraa43TNDFCuvPmKJfO95lnhy4'
bot = telebot.TeleBot("1301173613:AAHpTOmYPAraa43TNDFCuvPmKJfO95lnhy4")

chat_all = '-1001485090186'
api_all = '1301173613:AAHpTOmYPAraa43TNDFCuvPmKJfO95lnhy4'

def erlog(*args):
	l = open("/var/www/html/erlog"+reg+".txt", 'a', encoding='windows-1251')		#в линуксе тут нужен полный путь до файла /var/www/html/NMU_log.txt
	print('Дата проверки '+str(datetime.today())[:-10], file=l)
	print(*args, file=l)
	l.close()

def log(*args):
    with open('nmu'+reg+'_log.txt', encoding='windows-1251') as fr:			# /var/www/prognoznmu.ru/nmu52_log.txt
        tm = fr.read()
    with open('nmu'+reg+'_log.txt', 'w', encoding='windows-1251') as fw:
        print('Дата проверки '+str(datetime.today())[:-10], file=fw)
        print(*args,  tm, sep='\n', file=fw)

if not os.path.isfile('re'+reg+'.xlsx'):
	wb = Workbook('re'+reg+'.xlsx')
	ws = wb.create_sheet("Проверка обновления")
	ws.append(["Дата", "Текст"])
	wb.save(os.path.dirname(os.path.abspath(__file__)) + "/re"+reg+".xlsx")
	wb = load_workbook('re'+reg+'.xlsx')
else:
	wb = load_workbook('re'+reg+'.xlsx')
ws = wb.active

while True:

	try:
		# time.sleep(random.randint(0, 300))
		ua = UserAgent()
		headers = {'User-Agent': ua.firefox}
		r = requests.get(url, headers=headers, timeout=15)
		r.encoding = 'windows-1251'
		soup = BeautifulSoup(r.text, 'html.parser')
		
		empty_tags = soup.findAll(lambda tag: not tag.contents or len(tag.get_text(strip=True)) <= 0)			# удаляем пустые теги
		[empty_tag.extract() for empty_tag in empty_tags]
		
		# while soup.find('div', {'class' : 'sub_menu'}):
			# soup.find('div', {'class' : 'sub_menu'}).decompose()
		# while soup.find('h1'):
			# soup.find('h1').decompose()

		st = soup.find('div', {'class' : 'text2'}).find('table')				#извлекаем текст из тела страницы

		st_c = re.sub(r'\<[^>]*\>', '', str(st))				# убираем всё, что в теговых скобках, оставляем только текст
		st_clean = re.sub('[\r\t\n]', '', st_c)
		st_clean = re.sub('Подробнее', '', st_clean)
		print(st_clean)
		now = datetime.today()
		
		# if ws['B2'].value != st_clean:				# проверка обновления прогноза
			# #print('Произошло обновление прогноза')
			# ws['A2'].value = str(datetime.today())[:-10]
			# ws['B2'].value = st_clean
			# ws.append([str(datetime.today())[:-10], st_clean])
			# wb.save(os.path.dirname(os.path.abspath(__file__)) + "/re"+reg+".xlsx")	#записываем новый прогноз в таблицу
			
			# bot.send_message(chatid, st_clean+'\n'+'Источник: [Брянский ЦГМС]('+url+')', parse_mode="Markdown", disable_web_page_preview="true")

			# bot.send_message(chat_all, 'Регион '+reg+' (Брянск)'+'\n'+st_clean+'\n'+'Источник: [Брянский ЦГМС]('+url+')',
			 # parse_mode="Markdown", disable_web_page_preview="true")


			# log('- '+st_clean+"\n")						# логируем извлеченную чистую строку

		# elif now.hour >= vecher and (now.month != int(ws['A2'].value[5:7]) or now.day != int(ws['A2'].value[8:10])): #не было обновления сегодня и наступил вечер
			# ws['A2'].value = str(datetime.today())[:-10]	#обновляем дату
			# ws.append([str(datetime.today())[:-10], 'Прогноз НМУ не изменился'])
			# wb.save(os.path.dirname(os.path.abspath(__file__)) + "/re"+reg+".xlsx")	#фиксируем отсутствие изменений в таблице
			
			# bot.send_message(chatid, 'Прогноз НМУ не изменился'+'\n'+'Источник: [Брянский ЦГМС]('+url+')', parse_mode="Markdown", disable_web_page_preview="true")

			# bot.send_message(chat_all, 'Регион '+reg+' (Брянск)'+'\n'+'Прогноз НМУ не изменился'+'\n'+'Источник: [Брянский ЦГМС]('+url+')',
			 # parse_mode="Markdown", disable_web_page_preview="true")

			
			# log('- Прогноз НМУ не изменился'+"\n")						# логируем строку
		
		
	except requests.ConnectionError as e:
		erlog(" Упс!! Ошибка соединения. Проверьте подключение к Интернет. Technical Details given below.")
		erlog(str(e))
		time.sleep(300)					#пауза 5 мин
		continue						#повторяем запрос
	except requests.Timeout as e:
		erlog(" OOPS!! Timeout Error")
		erlog(str(e))
		time.sleep(300)					#пауза 5 мин
		continue						#повторяем запрос
	except requests.RequestException as e:
		erlog("OOPS!! General Error")
		erlog(str(e))
		time.sleep(300)					#пауза 5 мин
		continue						#повторяем запрос
	except KeyboardInterrupt:
		erlog(" Кто-то закрыл программу")
	
	break
